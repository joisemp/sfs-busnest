"""
Celery tasks for the services app.

This module defines asynchronous background tasks for the SFS Institutions system, including:
- Email sending and newsletter delivery
- Processing uploaded Excel files for routes, receipts, and buses
- Exporting tickets to Excel
- Generating student pass PDFs
- Bulk updating student groups
- Marking expired receipts

Each task handles logging, error handling, and user notification via the Notification model. Utility functions are provided for sending export emails and generating download links.

Tasks:
    count_task: Example task that logs numbers 1 to 10.
    send_newsletter: Example task that simulates sending newsletters.
    send_email_task: Sends an email asynchronously.
    process_uploaded_route_excel: Processes uploaded route Excel files and creates Route/Stop objects.
    process_uploaded_receipt_data_excel: Processes uploaded receipt Excel files and creates Receipt/StudentGroup objects.
    process_uploaded_bus_excel: Processes uploaded bus Excel files and creates/updates Bus objects.
    mark_expired_receipts: Marks receipts as expired after a set number of days.
    export_tickets_to_excel: Exports filtered tickets to an Excel file and emails the user a download link.
    generate_student_pass: Generates a PDF of student passes and emails the user a download link.
    bulk_update_student_groups_task: Bulk updates student groups for tickets based on preview data.

Utility Functions:
    send_export_email: Sends an email with a download link for an exported file.
"""

from celery import shared_task
from click import File
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
import openpyxl
from openpyxl.styles import Font
from io import BytesIO
from django.conf import settings
import logging, time, os
from django.core.mail import send_mail
from services.models import Organisation, Receipt, Stop, Route, Institution, Registration, StudentGroup, Ticket, ExportedFile, BusFile, Bus, Notification, StudentPassFile
from django.db import transaction, models, IntegrityError
from django.db.models import Q
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.utils.timezone import now
from datetime import timedelta, datetime
from django.conf import settings
from django.core.files.storage import default_storage
from django.utils import timezone
from django.utils.text import slugify
from uuid import uuid4
from django.core.files.base import ContentFile
from services.utils.utils import generate_ids_pdf  # Import from utils instead of views
from urllib.parse import urljoin
from django.db.utils import IntegrityError

User = get_user_model()

logger = logging.getLogger(__name__)


@shared_task(name='count_to_10')
def count_task():
    """
    Example Celery task that logs numbers 1 to 10 with a delay.
    """
    logger.info("Task started: Count to 10")
    for i in range(1, 11):
        logger.info(i)
        time.sleep(1)
    logger.info("Task compeleted!!!!")
    return "TASK DONE!!"

@shared_task(name='send_newsletter')
def send_newsletter():
    """
    Example Celery task that simulates sending newsletters.
    """
    logger.info("Task started : sending newsletters")
    for i in range(1, 11):
        logger.info(f'{i}. Newsletter')
        time.sleep(1)
    logger.info("All Newsletter sent!!")
    return 'COMPLETED!!!!!'

@shared_task(name='send_email_task')
def send_email_task(subject, message, recipient_list, from_email=None):
    """
    Sends an email asynchronously using Django's send_mail.
    Args:
        subject (str): Email subject.
        message (str): Email body.
        recipient_list (list): List of recipient email addresses.
        from_email (str, optional): Sender email address.
    Returns:
        str: Success message if sent, raises exception on failure.
    """
    try:
        logger.info(f"Starting email task: Sending email to {recipient_list}")
        send_mail(
            subject=subject,
            message=message,
            from_email=from_email,
            recipient_list=recipient_list,
            fail_silently=False,
        )
        logger.info(f"Email successfully sent to {recipient_list}")
        return "Email sent successfully"
    except Exception as e:
        logger.error(f"Failed to send email to {recipient_list}: {e}")
        raise


@shared_task(name='process_uploaded_route_excel')
def process_uploaded_route_excel(user_id, file_path, org_id, registration_id):
    """
    Processes an uploaded route Excel file, creating Route and Stop objects, and notifies the user.
    Args:
        user_id (int): ID of the user who uploaded the file.
        file_path (str): Path to the uploaded file.
        org_id (int): Organization ID.
        registration_id (int): Registration ID.
    """
    try:
        user = User.objects.get(id=user_id)
        notification = Notification.objects.create(
            user=user,
            action="Route Excel Processing",
            description="<p>We have started processing your uploaded Route Excel file.</p>",
            file_processing_task=True,
            type="info"
        )

        logger.info(f"Task Started: Processing file: {file_path}")

        # Determine file location (local vs cloud storage)
        if settings.DEBUG:
            full_path = os.path.join(settings.MEDIA_ROOT, file_path)  # Local file
            file = open(full_path, 'rb')  # Open the local file
        else:
            full_path = file_path  # Cloud storage path
            file = default_storage.open(full_path, 'rb')  # Open the file from cloud storage

        processed_routes = 0
        skipped_routes = []
        skipped_stops = []

        with transaction.atomic():
            # Fetch Organisation
            try:
                org = Organisation.objects.get(id=org_id)
                logger.info(f"Organisation fetched successfully: {org.name} (ID: {org_id})")
            except Organisation.DoesNotExist:
                notification.action = "Organisation Not Found"
                notification.description = "<p>The organisation linked to the file could not be found. Please check and try again.</p>"
                notification.type = "danger"
                notification.save()
                return

            # Fetch Registration
            try:
                registration_obj = Registration.objects.get(id=registration_id)
                logger.info(f"Registration fetched successfully: {registration_obj.name} (ID: {registration_id})")
            except Registration.DoesNotExist:
                notification.action = "Registration Not Found"
                notification.description = "<p>The registration linked to the file could not be found. Please check and try again.</p>"
                notification.type = "danger"
                notification.save()
                return

            # Open and process the Excel file
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                logger.info(f"Excel file opened successfully: {file_path}")
            except Exception as e:
                notification.action = "File Open Error"
                notification.description = "<p>We couldn't open the uploaded file. Please ensure it is a valid Excel file and try again.</p>"
                notification.type = "danger"
                notification.save()
                raise
            finally:
                file.close()  # Ensure the file is closed after processing

            # Extract headers (route names)
            headers = [cell.value for cell in sheet[1]]  # First row as headers
            logger.info(f"Extracted headers (route names): {headers}")

            for col_index, route_name in enumerate(headers, start=1):
                column_letter = openpyxl.utils.get_column_letter(col_index)  # Convert column index to Excel-style letter
                if not route_name:
                    skipped_routes.append((column_letter, "No route name provided"))
                    continue

                try:
                    route, created = Route.objects.get_or_create(
                        org=org,
                        registration=registration_obj,
                        name=route_name.strip()
                    )

                    if created:
                        logger.info(f"Created Route: {route.name} (ID: {route.id})")
                    else:
                        logger.info(f"Route already exists: {route.name} (ID: {route.id})")

                    # Iterate over stops in the column
                    for row_number, row in enumerate(sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index), start=1):
                        stop_name = row[0].value
                        if not stop_name:
                            skipped_stops.append((row_number, column_letter, "No stop name provided"))
                            continue

                        stop_name = stop_name.strip().upper()

                        Stop.objects.get_or_create(
                            org=org,
                            registration=registration_obj,
                            route=route,
                            name=stop_name
                        )

                    processed_routes += 1

                except Exception as e:
                    notification.action = "Processing Error"
                    notification.description = "<p>An error occurred while processing the file. Please try again later.</p>"
                    notification.type = "danger"
                    notification.save()
                    raise

            notification.action = "Route Excel Processed"
            notification.description = (
                f"<p>The Route Excel file has been processed successfully.</p>"
                f"<p>Routes added: {processed_routes}.</p>"
                f"<p>Routes skipped: {len(skipped_routes)}.</p>"
                f"<p>Stops skipped: {len(skipped_stops)}.</p>"
            )
            if skipped_routes:
                notification.description += "<p>Details of skipped routes:</p><ul>"
                for column_letter, reason in skipped_routes:
                    notification.description += f"<li>Column {column_letter}: {reason}</li>"
                notification.description += "</ul>"
            if skipped_stops:
                notification.description += "<p>Details of skipped stops:</p><ul>"
                for row_number, column_letter, reason in skipped_stops:
                    notification.description += f"<li>Row {row_number}, Column {column_letter}: {reason}</li>"
                notification.description += "</ul>"
            notification.type = "success"
            notification.save()

    except Exception as e:
        notification.action = "Unexpected Error"
        notification.description = "<p>An unexpected error occurred while processing the file. Please contact support if the issue persists.</p>"
        notification.type = "danger"
        notification.save()
        raise
    finally:
        logger.info("Task Ended: process_uploaded_route_excel")
        notification.save()


@shared_task(name='process_uploaded_receipt_data_excel')
def process_uploaded_receipt_data_excel(user_id, file_path, org_id, institution_id, reg_id):
    """
    Processes an uploaded receipt data Excel file, creating Receipt and StudentGroup objects, and notifies the user.
    Args:
        user_id (int): ID of the user who uploaded the file.
        file_path (str): Path to the uploaded file.
        org_id (int): Organization ID.
        institution_id (int): Institution ID.
        reg_id (int): Registration ID.
    """
    try:
        user = User.objects.get(id=user_id)
        notification = Notification.objects.create(
            user=user,
            action="Receipt Data Excel Processing",
            description="<p>Processing Receipt Data Excel file has started.</p>",
            file_processing_task=True,
            type="info"
        )

        logger.info(f"Task Started: Processing file: {file_path}")

        # Determine file location (local vs cloud storage)
        if settings.DEBUG:
            full_path = os.path.join(settings.MEDIA_ROOT, file_path)  # Local file
            file = open(full_path, 'rb')  # Open the local file
        else:
            full_path = file_path  # Cloud storage path
            file = default_storage.open(full_path, 'rb')  # Open the file from cloud storage

        processed_count = 0
        skipped_rows = []

        valid_classes = [
            'LKG', 'UKG', 'PRE-KG', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12',
            'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII'
        ]
        valid_sections = [chr(i).upper() for i in range(65, 91)] + ['teaching', 'non-teaching', 'Teaching', 'Non-Teaching']
        
        # Fetch Organisation
        try:
            org = Organisation.objects.get(id=org_id)
            logger.info(f"Organisation fetched successfully: {org.name} (ID: {org_id})")
        except Organisation.DoesNotExist:
            error_message = "Organisation does not exist."
            logger.error(error_message)
            notification.action = error_message
            notification.description = f"<p>Organisation with ID {org_id} does not exist.</p>"
            notification.type = "danger"
            notification.save()
            return

        # Fetch Institution
        try:
            institution = Institution.objects.get(id=institution_id)
            logger.info(f"Institution fetched successfully: {institution.name} (ID: {institution_id})")
        except Institution.DoesNotExist:
            error_message = "Institution does not exist."
            logger.error(error_message)
            notification.action = error_message
            notification.description = f"<p>Institution with ID {institution_id} does not exist.</p>"
            notification.type = "danger"
            notification.save()
            return

        # Fetch Registration
        try:
            registration = Registration.objects.get(id=reg_id)
            logger.info(f"Registration fetched successfully: {registration.name} (ID: {reg_id})")
        except Registration.DoesNotExist:
            error_message = "Registration does not exist."
            logger.error(error_message)
            notification.action = error_message
            notification.description = f"<p>Registration with ID {reg_id} does not exist.</p>"
            notification.type = "danger"
            notification.save()
            return
        
        # Open Excel file
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            logger.info(f"Excel file opened successfully: {file_path}")
        except Exception as e:
            error_message = "Failed to open the Excel file."
            logger.error(error_message)
            notification.action = error_message
            notification.description = f"<p>Failed to open the Excel file: {e}</p>"
            notification.type = "danger"
            notification.save()
            raise
        finally:
            file.close()  # Ensure the file is closed after processing
        
        #-----------------------------

        with transaction.atomic():
            # Validate headers
            headers = [str(cell.value).strip().lower() for cell in sheet[1]]
            expected_headers = ['receipt id', 'student id', 'class', 'section']
            if headers != expected_headers:
                error_message = "Invalid headers in the Receipt Data Excel file."
                logger.error(error_message)
                notification.action = error_message
                notification.description = f"<p>Expected headers: {expected_headers}</p><p>Found headers: {headers}</p>"
                notification.type = "danger"
                notification.save()
                return

            # Process rows
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                receipt_id, student_id, class_name, class_section = map(str, row)

                # Strip and validate class and section
                class_name = class_name.strip().upper().replace(" ", "")
                class_section = class_section.strip().upper().replace(" ", "")

                if not receipt_id or not student_id or class_name not in valid_classes or class_section not in valid_sections:
                    warning_message = f"Row {row_number}: Skipping invalid or incomplete row: {row}"
                    logger.warning(warning_message)
                    skipped_rows.append((row_number, row, "Invalid or incomplete data"))
                    continue

                group_name = f"{class_name} - {class_section}"

                try:
                    with transaction.atomic():
                        # Create or Get StudentGroup
                        try:
                            student_group, created = StudentGroup.objects.get_or_create(
                                org=org,
                                institution=institution,
                                name=group_name.upper(),
                            )
                            if created:
                                logger.info(f"Student Group created: {student_group.name} (ID: {student_group.id})")
                            else:
                                logger.info(f"Student Group already exists: {student_group.name} (ID: {student_group.id})")
                        except IntegrityError:
                            logger.warning(f"Race condition detected while creating Student Group: {group_name.upper()}. Retrying...")
                            student_group = StudentGroup.objects.get(
                                org=org,
                                institution=institution,
                                name=group_name.upper(),
                            )

                        try:
                            # Create Receipt
                            Receipt.objects.create(
                                org=org,
                                institution=institution,
                                registration=registration,
                                receipt_id=receipt_id.strip(),
                                student_id=student_id.strip().upper(),
                                student_group=student_group,
                            )
                            logger.info(f"Row {row_number}: Receipt created - Receipt ID: {receipt_id}, Student ID: {student_id}")
                            processed_count += 1

                        except IntegrityError as e:
                            # Handle duplicate receipts
                            error_message = f"Row {row_number}: Duplicate Receipt - Receipt ID: {receipt_id}, Student ID: {student_id}"
                            logger.warning(error_message)
                            skipped_rows.append((row_number, row, "Duplicate receipt"))
                            continue

                        except Exception as e:
                            # Handle unexpected errors
                            error_message = f"Row {row_number}: Unexpected error while creating Receipt - {e}"
                            logger.error(error_message)
                            skipped_rows.append((row_number, row, str(e)))
                            continue

                except Exception as e:
                    error_message = f"Row {row_number}: Error processing receipt - {e}"
                    logger.error(error_message)
                    skipped_rows.append((row_number, row, str(e)))
                    continue

            success_message = "Receipt Data Excel file processed successfully."
            logger.info(success_message)
            notification.action = success_message
            notification.description = (
                f"<p>Total processed rows: {processed_count}</p>"
                f"<p>Total skipped rows: {len(skipped_rows)}</p>"
            )
            if skipped_rows:
                notification.description += "<p>Details of skipped rows:</p><ul>"
                for row_number, row, reason in skipped_rows:
                    notification.description += f"<li>Row {row_number}: {row} - {reason}</li>"
                notification.description += "</ul>"
            notification.type = "success"
            notification.save()

        #----------------------------
        
    except Exception as e:
        error_message = "Error while processing Receipt Data Excel file."
        logger.error(error_message)
        notification.action = error_message
        notification.description = f"<p>{e}</p>"
        notification.type = "danger"
        notification.save()
        raise
    finally:
        logger.info("Task Ended: process_uploaded_receipt_data_excel")
        notification.save()
        

@shared_task(name='process_uploaded_bus_excel')
def process_uploaded_bus_excel(user_id, file_path, org_id):
    """
    Processes an uploaded bus Excel file, creating or updating Bus objects, and notifies the user.
    Args:
        user_id (int): ID of the user who uploaded the file.
        file_path (str): Path to the uploaded file.
        org_id (int): Organization ID.
    """
    try:
        user = User.objects.get(id=user_id)
        notification = Notification.objects.create(
            user=user,
            action="Bus Excel Processing",
            description="<p>Processing Bus Excel file has started.</p>",
            file_processing_task=True,
            type="info"
        )
        
        logger.info(f"Task Started: Processing file: {file_path}")

        # Determine file location (local vs cloud storage)
        if settings.DEBUG:
            full_path = os.path.join(settings.MEDIA_ROOT, file_path)  # Local file
            file = open(full_path, 'rb')  # Open the local file
        else:
            full_path = file_path  # Cloud storage path
            file = default_storage.open(full_path, 'rb')  # Open the file from cloud storage

        processed_count = 0
        skipped_rows = []

        with transaction.atomic():
            # Fetch organisation
            try:
                org = Organisation.objects.get(id=org_id)
                logger.info(f"Organisation fetched successfully: {org.name} (ID: {org_id})")
            except Organisation.DoesNotExist:
                error_message = "Organisation does not exist."
                logger.error(error_message)
                notification.action = error_message
                notification.description = f"<p>Organisation with ID {org_id} does not exist.</p>"
                notification.type = "danger"
                notification.save()
                return

            # Fetch BusFile entry
            try:
                bus_file = BusFile.objects.get(file=file_path)
                logger.info(f"BusFile entry fetched: {bus_file.name}")
            except BusFile.DoesNotExist:
                error_message = "BusFile does not exist."
                logger.error(error_message)
                notification.action = error_message
                notification.description = f"<p>BusFile with path {file_path} does not exist.</p>"
                notification.type = "danger"
                notification.save()
                return

            # Open Excel file correctly
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                logger.info(f"Excel file opened successfully: {file_path}")
            except Exception as e:
                error_message = "Failed to open the Excel file."
                logger.error(error_message)
                notification.action = error_message
                notification.description = f"<p>Failed to open the Excel file: {e}</p>"
                notification.type = "danger"
                notification.save()
                raise
            finally:
                file.close()  # Close file after processing

            # Validate headers
            headers = [cell.value for cell in sheet[1]]
            expected_headers = ["Registration", "Driver", "Capacity"]
            if headers != expected_headers:
                error_message = "Invalid headers in the Bus Excel file."
                logger.error(error_message)
                notification.action = error_message
                notification.description = f"<p>Expected headers: {expected_headers}</p><p>Found headers: {headers}</p>"
                notification.type = "danger"
                notification.save()
                return

            # Process rows
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                registration, driver, capacity = row

                if not registration or not driver or not capacity:
                    warning_message = f"Row {row_number}: Skipping incomplete row: {row}"
                    logger.warning(warning_message)
                    skipped_rows.append((row_number, row))
                    continue

                try:
                    bus, created = Bus.objects.get_or_create(
                        org=org,
                        registration_no=registration.strip(),
                        defaults={
                            'driver': driver.strip(),
                            'capacity': int(capacity)
                        }
                    )

                    if created:
                        logger.info(f"Row {row_number}: Bus created - Registration: {registration}, Driver: {driver}, Capacity: {capacity}")
                    else:
                        bus.driver = driver.strip()
                        bus.capacity = int(capacity)
                        bus.save()
                        logger.info(f"Row {row_number}: Bus updated - Registration: {registration}, Driver: {driver}, Capacity: {capacity}")

                    processed_count += 1

                except Exception as e:
                    error_message = f"Row {row_number}: Error processing bus {registration}: {e}"
                    logger.error(error_message)
                    notification.action = error_message
                    notification.description += f"<p>{error_message}</p>"
                    notification.type = "danger"
                    notification.save()
                    raise

            bus_file.added = True
            bus_file.save()

            success_message = "Bus file processed successfully."
            logger.info(success_message)
            notification.action = success_message
            notification.description = (
                f"<p>All rows in the Bus Excel file have been processed successfully.</p>"
                f"<p>Total processed rows: {processed_count}</p>"
                f"<p>Total skipped rows: {len(skipped_rows)}</p>"
            )
            if skipped_rows:
                notification.description += "<p>Details of skipped rows:</p><ul>"
                for row_number, row in skipped_rows:
                    notification.description += f"<li>Row {row_number}: {row}</li>"
                notification.description += "</ul>"
            notification.type = "success"
            notification.save()

    except Exception as e:
        error_message = "Error while processing Excel file."
        logger.error(error_message)
        notification.action = error_message
        notification.description += f"<p>{e}</p>"
        notification.type = "danger"
        notification.save()
        raise
    finally:
        logger.info("Task Ended: process_uploaded_bus_excel")
        notification.save()


@shared_task
def mark_expired_receipts():
    """
    Marks receipts as expired if they are older than a set number of days.
    Returns:
        str: Log message indicating how many receipts were marked as expired.
    """
    expiry_days = 7 
    expiry_date = now() - timedelta(days=expiry_days)

    updated_count = Receipt.objects.filter(created_at__lt=expiry_date, is_expired=False).update(is_expired=True)

    log_message = f"Marked {updated_count} receipts as expired."
    logger.info(log_message)
    return log_message


def send_export_email(user, exported_file):
    """
    Sends an email to the user with a download link for the exported file.
    Args:
        user (User): The user to notify.
        exported_file (ExportedFile): The exported file object.
    """
    from urllib.parse import urljoin

    download_url = reverse('central_admin:exported_file_download', kwargs={'slug': exported_file.slug})
    full_url = urljoin(settings.SITE_URL, download_url)
    email_subject = 'Your Ticket Export is Ready'
    email_message = (
        f"Hello {user.profile.first_name} {user.profile.last_name},\n\n"
        f"Your ticket export is ready. You can download the file using the following link:\n"
        f"{full_url}\n\n"
        f"Best regards,\nYour Team"
    )
    send_mail(
        email_subject,
        email_message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email]
    )

@shared_task(name='export_tickets_to_excel')
def export_tickets_to_excel(user_id, registration_slug, search_term='', filters=None):
    """
    Exports filtered tickets to an Excel file and emails the user a download link.
    Args:
        user_id (int): ID of the user requesting the export.
        registration_slug (str): Registration slug.
        search_term (str, optional): Search term for filtering tickets.
        filters (dict, optional): Additional filters for the queryset.
    Returns:
        str: Success message after export and email.
    """
    user = User.objects.get(id=user_id)
    registration = get_object_or_404(Registration, slug=registration_slug)

    queryset = Ticket.objects.filter(org=user.profile.org, registration=registration).order_by('-created_at')

    if search_term:
        queryset = queryset.filter(
            Q(student_name__icontains=search_term) |
            Q(student_email__icontains=search_term) |
            Q(student_id__icontains=search_term) |
            Q(contact_no__icontains=search_term) |
            Q(alternative_contact_no__icontains=search_term)
        )

    if filters:
        if filters.get('institution'):
            queryset = queryset.filter(institution_id=filters['institution'])
        if filters.get('pickup_points'):
            queryset = queryset.filter(pickup_point_id__in=filters['pickup_points'])
        if filters.get('drop_points'):
            queryset = queryset.filter(drop_point_id__in=filters['drop_points'])
        if filters.get('schedule'):
            queryset = queryset.filter(schedule_id=filters['schedule'])
        if filters.get('pickup_buses'):
            queryset = queryset.filter(pickup_bus_record_id__in=filters['pickup_buses'])
        if filters.get('drop_buses'):
            queryset = queryset.filter(drop_bus_record_id__in=filters['drop_buses'])
        if filters.get('student_group'):
            queryset = queryset.filter(student_group_id=filters['student_group'])
        if filters.get('pickup_schedule'):
            queryset = queryset.filter(pickup_schedule_id=filters['pickup_schedule'])
        if filters.get('drop_schedule'):
            queryset = queryset.filter(drop_schedule_id=filters['drop_schedule'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    headers = [
        'TICKET ID', 'STUDENT ID', 'STUDENT NAME', 'CLASS', 'SECTION', 'STUDENT EMAIL', 'CONTACT NO', 
        'ALTERNATIVE NO', 'PICKUP POINT', 'DROP POINT', 'PICKUP BUS', 'DROP BUS', 
        'PICKUP SCHEDULE', 'DROP SCHEDULE', 'INSTITUTION', 'STATUS', 'CREATED AT'
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for ticket in queryset:
        # Safely split student group name into class and section
        student_group_name = str(ticket.student_group.name)
        if '-' in student_group_name:
            std_class, section = student_group_name.split('-', 1)  # Split into at most 2 parts
        else:
            std_class, section = student_group_name, ''  # Default section to an empty string if no hyphen

        ws.append([
            ticket.ticket_id,
            ticket.student_id,
            ticket.student_name.upper(),
            std_class.strip().upper(),
            section.strip().upper(),
            ticket.student_email,
            ticket.contact_no.upper(),
            ticket.alternative_contact_no.upper(),
            (ticket.pickup_point.name.upper() if ticket.pickup_point else '-----'),
            (ticket.drop_point.name.upper() if ticket.drop_point else '-----'),
            (ticket.pickup_bus_record.label.upper() if ticket.pickup_bus_record else '-----'),
            (ticket.drop_bus_record.label.upper() if ticket.drop_bus_record else '-----'),
            (ticket.pickup_schedule.name.upper() if ticket.pickup_schedule else '-----'),
            (ticket.drop_schedule.name.upper() if ticket.drop_schedule else '-----'),
            (ticket.institution.name.upper() if ticket.institution else '-----'),
            'CONFIRMED' if ticket.status else 'PENDING',
            ticket.created_at.strftime('%Y-%m-%d %H:%M:%S').upper()
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    unique_slug = slugify(f"{registration_slug}-{uuid4()}")
    exported_file = ExportedFile.objects.create(
        user=user,
        file=ContentFile(file_stream.read(), f"{registration_slug}_export.xlsx"),
        slug=unique_slug
    )

    send_export_email(user, exported_file)
    return f"Excel export completed for {user.profile.first_name} {user.profile.last_name} ({user.email})"


@shared_task(name='generate_student_pass')
def generate_student_pass(user_id, registration_slug, filters=None):
    """
    Generates a PDF of student passes for filtered tickets and emails the user a download link.
    Args:
        user_id (int): ID of the user requesting the generation.
        registration_slug (str): Registration slug.
        filters (dict, optional): Additional filters for the queryset.
    Returns:
        str: Success message after generation and email.
    """
    user = User.objects.get(id=user_id)
    registration = get_object_or_404(Registration, slug=registration_slug)

    # Base queryset
    queryset = Ticket.objects.filter(org=user.profile.org, registration=registration).order_by('-created_at')

    # Apply filters
    if filters:
        if filters.get('start_date') or filters.get('end_date'):
            try:
                end_date_str = filters.get('end_date') or now().date().isoformat()
                start_date_str = filters.get('start_date') or end_date_str

                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()

                queryset = queryset.filter(
                    created_at__date__range=[start_date, end_date]
                )
            except ValueError:
                logger.error("Invalid date format in filters. Ensure dates are in 'YYYY-MM-DD' format.")
                raise ValueError("Invalid date format in filters. Ensure dates are in 'YYYY-MM-DD' format.")

        if filters.get('institution'):
            queryset = queryset.filter(institution__slug=filters['institution'])
        if filters.get('ticket_type'):
            queryset = queryset.filter(ticket_type=filters['ticket_type'])
        if filters.get('student_group'):
            queryset = queryset.filter(student_group_id=filters['student_group'])

    # Generate the PDF using the filtered queryset
    students = queryset.order_by('institution__name', 'student_group__name').values(
        'student_name', 'pickup_bus_record__label',
        'drop_bus_record__label', 'institution__name',
        'student_id', 'ticket_id', 'pickup_schedule__name', 'drop_schedule__name',
        'student_group__name'
    )
    buffer = generate_ids_pdf(students)

    # Create a StudentPassFile instance
    unique_slug = slugify(f"{registration_slug}-{uuid4()}")
    student_pass_file = StudentPassFile.objects.create(
        user=user,
        file=ContentFile(buffer.getvalue(), f"{registration_slug}_student_passes.pdf"),
        slug=unique_slug
    )

    
    download_url = reverse('central_admin:student_pass_file_download', kwargs={'slug': student_pass_file.slug})
    full_url = urljoin(settings.SITE_URL, download_url)
    email_subject = 'Your Student Pass is Ready'
    email_message = (
        f"Hello {user.profile.first_name} {user.profile.last_name},\n\n"
        f"Your student pass is ready. You can download the file using the following link:\n"
        f"{full_url}\n\n"
        f"Best regards,\nSFS BusNest Team"
    )
    send_mail(
        email_subject,
        email_message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email]
    )

    return f"Student pass generation completed for {user.profile.first_name} {user.profile.last_name} ({user.email})"


@shared_task
def bulk_update_student_groups_task(user_id, institution_id, preview_data):
    """
    Bulk updates student groups for tickets based on preview data.
    Args:
        user_id (int): ID of the user initiating the update.
        institution_id (int): Institution ID.
        preview_data (list): List of dictionaries with student_id and new_group.
    """
    from services.models import Ticket, StudentGroup, Institution
    user = User.objects.get(id=user_id)
    institution = Institution.objects.get(id=institution_id)
    org = institution.org
    message = "The Student Group update process has been initiated. You will be notified once the process is complete."
    notification = Notification.objects.create(
        user=user,
        action="Bulk Student Group Update",
        description=f"<p>{message}</p>",
        type="info",
        file_processing_task=True
    )
    logger.info(f"Bulk Student Group Update Task started for user: {user.email} in institution: {institution.name}")
    for entry in preview_data:
        student_id = entry['student_id']
        new_group_name = entry['new_group']
        try:
            ticket = Ticket.objects.get(student_id=student_id, institution=institution)
            student_group, _ = StudentGroup.objects.get_or_create(
                org=org, institution=institution, name=new_group_name
            )
            ticket.student_group = student_group
            ticket.save()
            logger.info(f"Updated Student Group for Ticket ID: {ticket.ticket_id}, Student ID: {student_id} to {new_group_name}")
        except Ticket.DoesNotExist:
            logger.warning(f"Ticket not found for Student ID: {student_id} in institution: {institution.name}")
            continue
    # Optionally: create a Notification for the user
    notification.action = "Bulk Student Group Update Completed"
    notification.description = "<p>The Student Group update process has been completed successfully.</p>"
    notification.type = "success"
    notification.save()

