"""
Views for Institution Admin functionalities in the SFS Institutions system.

This module contains Django class-based views for managing registrations, tickets, receipts, student groups, bus searches, bus requests, and bulk updates for institution administrators. It includes list, create, update, and delete views, as well as custom logic for searching and updating bus and student group information.

Classes:
    RegistrationListView: Lists registrations for the current user's organization.
    TicketListView: Lists and filters tickets for a registration and institution.
    TicketUpdateView: Updates ticket details and related receipt institution.
    TicketDeleteView: Deletes a ticket.
    ReceiptListView: Lists receipts for a registration.
    ReceiptDataFileUploadView: Handles uploading and processing of receipt data Excel files.
    ReceiptCreateView: Creates a new receipt.
    ReceiptDeleteView: Deletes a receipt.
    StudentGroupListView: Lists student groups for a registration.
    StudentGroupCreateView: Creates a new student group.
    StudentGroupUpdateView: Updates a student group.
    StudentGroupDeleteView: Deletes a student group.
    BusSearchFormView: Handles bus search form for a registration.
    BusSearchResultsView: Displays bus search results based on session criteria.
    TicketExportView: Triggers export of tickets to Excel via Celery.
    StopSelectFormView: Handles stop selection form for students.
    SelectScheduleGroupView: Handles schedule group selection for a ticket.
    BusSearchResultsView: (students) Displays filtered bus records for a stop and schedule.
    UpdateBusInfoView: Updates ticket's bus info and trip booking counts.
    BusRequestListView: Lists bus requests for a registration.
    BusRequestOpenListView: Lists open bus requests.
    BusRequestClosedListView: Lists closed bus requests.
    BusRequestDeleteView: Deletes a bus request.
    BusRequestStatusUpdateView: Updates the status of a bus request and adds comments.
    BusRequestCommentView: Adds a comment to a bus request.
    BulkStudentGroupUpdateView: Handles bulk update of student groups via Excel upload.
    BulkStudentGroupUpdateConfirmView: Confirms and processes bulk student group updates.

Each view is protected by login and institution admin access mixins, and many use Celery tasks for background processing of long-running operations.
"""

import threading
from urllib.parse import urlencode
from django.db import transaction
from django.http import Http404, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import ListView, CreateView, DeleteView, UpdateView, FormView, View, DetailView
from django.urls import reverse, reverse_lazy
from services.forms.central_admin import BusRequestCommentForm
from services.forms.students import StopSelectForm
from services.models import Bus, BusRecord, BusRequest, BusRequestComment, Registration, Receipt, ScheduleGroup, Stop, StudentGroup, Ticket, Schedule, ReceiptFile, Trip, BusReservationRequest
from services.forms.institution_admin import ReceiptForm, StudentGroupForm, TicketForm, BusSearchForm, BulkStudentGroupUpdateForm, BusReservationRequestForm
from config.mixins.access_mixin import InsitutionAdminOnlyAccessMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Count
from django.template.loader import render_to_string
from services.tasks import process_uploaded_receipt_data_excel, export_tickets_to_excel, bulk_update_student_groups_task
from services.utils.utils import get_filtered_bus_records
import openpyxl
from django.contrib import messages

class RegistrationListView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, ListView):
    """
    View to list all registrations for the current user's organization.
    """
    model = Registration
    template_name = 'institution_admin/registration_list.html'
    context_object_name = 'registrations'
    
    def get_queryset(self):
        """
        Returns queryset of registrations filtered by the user's organization.
        """
        queryset = Registration.objects.filter(org=self.request.user.profile.org)
        return queryset
    

class TicketListView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, ListView):
    """
    View to list and filter tickets for a registration and institution.
    """
    model = Ticket
    template_name = 'institution_admin/ticket_list.html'
    context_object_name = 'tickets'
    paginate_by = 15
    
    def get_queryset(self):
        """
        Returns queryset of tickets filtered by registration, institution, and optional GET parameters.
        """
        # Get registration based on slug
        registration_slug = self.kwargs.get('registration_slug')
        self.registration = get_object_or_404(Registration, slug=registration_slug)
        
        # Base queryset filtered by registration and institution, excluding terminated tickets
        queryset = Ticket.objects.filter(
            registration=self.registration, 
            institution=self.request.user.profile.institution,
            is_terminated=False
        ).order_by('-created_at')
        
        # Apply filters based on GET parameters
        pickup_points = self.request.GET.getlist('pickup_point')
        drop_points = self.request.GET.getlist('drop_point')
        schedule = self.request.GET.get('schedule')
        student_group = self.request.GET.get('student_group')
        filters = False  # Default no filters applied
        
        self.search_term = self.request.GET.get('search', '')
        
        if self.search_term:
            queryset = Ticket.objects.filter(
                Q(student_name__icontains=self.search_term) |
                Q(student_email__icontains=self.search_term) |
                Q(student_id__icontains=self.search_term) |
                Q(contact_no__icontains=self.search_term) |
                Q(alternative_contact_no__icontains=self.search_term),
                registration=self.registration,
                institution=self.request.user.profile.institution,
                is_terminated=False
            )

        # Apply filters based on GET parameters and update the filters flag
        if pickup_points and not pickup_points == ['']:
            queryset = queryset.filter(pickup_point_id__in=pickup_points)
            filters = True
        if drop_points and not drop_points == ['']:
            queryset = queryset.filter(drop_point_id__in=drop_points)
            filters = True
        if schedule:
            queryset = queryset.filter(schedule_id=schedule)
            filters = True
        if student_group:
            queryset = queryset.filter(student_group_id=student_group)
            filters = True
        
        # Pass the filters flag to context (done in get_context_data)
        self.filters = filters  # Store in the instance for later access

        return queryset
    
    def get_context_data(self, **kwargs):
        """
        Adds filter status and filter options to the context for the template.
        """
        # Get default context from parent
        context = super().get_context_data(**kwargs)
        
        # Add the filter status to the context
        context['filters'] = self.filters  # Pass the filters flag to the template
        
        # Add the filter options to the context
        context['registration'] = self.registration
        context['pickup_points'] = Stop.objects.filter(
            org = self.request.user.profile.org
        )
        context['drop_points'] = Stop.objects.filter(
            org = self.request.user.profile.org
        )
        context['schedules'] = Schedule.objects.filter(
            org = self.request.user.profile.org
        )
        context['student_groups'] = StudentGroup.objects.filter(
            org = self.request.user.profile.org,
            institution = self.request.user.profile.institution
        ).order_by('name')
        context['search_term'] = self.search_term
        
        # Preserve query parameters for pagination
        query_dict = self.request.GET.copy()
        if 'page' in query_dict:
            query_dict.pop('page')
        context['query_params'] = query_dict.urlencode()

        return context


class TicketUpdateView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, UpdateView):
    """
    View to update ticket details and ensure receipt institution matches ticket institution.
    """
    model = Ticket
    form_class = TicketForm
    template_name = 'institution_admin/ticket_update.html'
    slug_url_kwarg = 'ticket_slug'

    def get_form_kwargs(self):
        """
        Add institution to form kwargs to filter student groups.
        """
        kwargs = super().get_form_kwargs()
        kwargs['institution'] = self.request.user.profile.institution
        return kwargs

    def form_valid(self, form):
        """
        Updates the ticket and its related receipt's institution if needed.
        """
        ticket = form.save()
        
        if ticket.institution != ticket.recipt.institution:
            ticket.recipt.institution = ticket.institution
            ticket.recipt.save()
            
        return super().form_valid(form)
    
    def get_success_url(self):
        """
        Returns the URL to redirect to after a successful update.
        """
        return reverse(
            'institution_admin:ticket_list', 
            kwargs={'registration_slug': self.kwargs['registration_slug']}
            )


class TicketDeleteView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, View):
    """
    View to soft delete (terminate) a ticket.
    """
    def get(self, request, registration_slug, ticket_slug):
        """
        Display confirmation page for ticket termination.
        """
        ticket = get_object_or_404(Ticket, slug=ticket_slug, registration__slug=registration_slug)
        return render(request, 'institution_admin/ticket_confirm_delete.html', {'object': ticket})
    
    def post(self, request, registration_slug, ticket_slug):
        """
        Soft delete the ticket by marking it as terminated.
        """
        ticket = get_object_or_404(Ticket, slug=ticket_slug, registration__slug=registration_slug)
        ticket.terminate()
        messages.success(request, f'Ticket for {ticket.student_name} has been terminated successfully.')
        return redirect('institution_admin:ticket_list', registration_slug=registration_slug)


class ReceiptListView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, ListView):
    """
    View to list all receipts for a registration.
    """
    model = Receipt
    template_name = 'institution_admin/receipt_list.html'
    context_object_name = 'receipts'
    paginate_by = 30

    def get_queryset(self):
        """
        Returns queryset of receipts filtered by registration, organization, and institution.
        """
        registration_slug = self.kwargs.get('registration_slug')
        self.registration = get_object_or_404(Registration, slug=registration_slug)
        queryset = Receipt.objects.filter(
            org=self.request.user.profile.org,
            institution=self.request.user.profile.institution,
            registration=self.registration
        ).order_by('-created_at')
        return queryset

    def get_context_data(self, **kwargs):
        """
        Adds registration to the context for the template.
        """
        context = super().get_context_data(**kwargs)
        context['registration'] = self.registration  # Ensure registration is passed to the template
        return context
    

class ReceiptDataFileUploadView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, CreateView):
    """
    View to handle uploading and background processing of receipt data Excel files.
    """
    model = ReceiptFile
    fields = ['registration', 'file']
    template_name = 'institution_admin/receipt_file_upload.html'
    
    def form_valid(self, form):
        """
        Saves the uploaded file and triggers background processing via Celery.
        """
        receipt_data_file = form.save(commit=False)
        user = self.request.user
        receipt_data_file.org = user.profile.org
        receipt_data_file.institution = user.profile.institution
        receipt_data_file.save()
        process_uploaded_receipt_data_excel.delay(
            self.request.user.id,
            receipt_data_file.file.name,
            user.profile.org.id,
            user.profile.institution.id,
            receipt_data_file.registration.id
        )
        return redirect(
            reverse(
                'institution_admin:receipt_list',
                kwargs={'registration_slug': receipt_data_file.registration.slug}
            )
        )
    
    
class ReceiptCreateView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, CreateView):
    """
    View to create a new receipt.
    """
    template_name = 'institution_admin/receipt_create.html'
    model = Receipt
    form_class = ReceiptForm
    
    def get_form_kwargs(self):
        """
        Add institution to form kwargs to filter student groups.
        """
        kwargs = super().get_form_kwargs()
        kwargs['institution'] = self.request.user.profile.institution
        return kwargs
    
    def form_valid(self, form):
        """
        Saves the new receipt and redirects to the receipt list.
        """
        receipt = form.save(commit=False)
        user = self.request.user
        receipt.org = user.profile.org
        receipt.institution = user.profile.institution
        receipt.save()
        return redirect(
            reverse(
                'institution_admin:receipt_list',
                kwargs={'registration_slug': receipt.registration.slug}
            )
        )
    

class ReceiptDeleteView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, DeleteView):
    """
    View to delete a receipt.
    """
    model = Receipt
    template_name = 'institution_admin/receipt_confirm_delete.html'
    slug_url_kwarg = 'receipt_slug'

    def get_success_url(self):
        """
        Returns the URL to redirect to after a successful delete.
        """
        return reverse(
            'institution_admin:receipt_list',
            kwargs={'registration_slug': self.object.registration.slug}
        )
    
    
class StudentGroupListView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, ListView):
    """
    View to list all student groups for a registration.
    """
    model = StudentGroup
    template_name = 'institution_admin/student_group_list.html'
    context_object_name = 'student_groups'  
    
    def get_queryset(self):
        """
        Returns queryset of student groups filtered by organization and institution.
        """
        registration_slug = self.kwargs.get('registration_slug')
        self.registration = get_object_or_404(Registration, slug=registration_slug)
        queryset = StudentGroup.objects.filter(
            org=self.request.user.profile.org,
            institution=self.request.user.profile.institution
        ).order_by('name')
        return queryset
    
    def get_context_data(self, **kwargs):
        """
        Adds registration to the context for the template.
        """
        context = super().get_context_data(**kwargs)
        context['registration'] = self.registration  # Ensure registration is passed to the template
        return context
    
    
class StudentGroupCreateView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, CreateView):
    """
    View to create a new student group.
    """
    template_name = 'institution_admin/student_group_create.html'
    model = StudentGroup
    form_class = StudentGroupForm

    def form_valid(self, form):
        """
        Saves the new student group and redirects to the student group list.
        """
        student_group = form.save(commit=False)
        user = self.request.user
        student_group.org = user.profile.org
        student_group.institution = user.profile.institution
        student_group.save()
        return redirect(
            reverse(
                'institution_admin:student_group_list',
                kwargs={'registration_slug': self.kwargs['registration_slug']}
            )
        )


class StudentGroupUpdateView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, UpdateView):
    """
    View to update a student group.
    """
    model = StudentGroup
    form_class = StudentGroupForm
    template_name = 'institution_admin/student_group_update.html'
    slug_url_kwarg = 'student_group_slug'

    def get_success_url(self):
        """
        Returns the URL to redirect to after a successful update.
        """
        return reverse(
            'institution_admin:student_group_list',
            kwargs={'registration_slug': self.kwargs['registration_slug']}
        )

    def form_valid(self, form):
        """
        Handles the form submission for updating a student group.
        """
        return super().form_valid(form)
    
    
class StudentGroupDeleteView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, DeleteView):
    """
    View to delete a student group.
    """
    model = StudentGroup
    template_name = 'institution_admin/student_group_confirm_delete.html'
    slug_url_kwarg = 'student_group_slug'

    def get_context_data(self, **kwargs):
        """
        Adds registration to the context for the template.
        """
        context = super().get_context_data(**kwargs)
        context['registration'] = get_object_or_404(Registration, slug=self.kwargs['registration_slug'])
        return context

    def get_success_url(self):
        """
        Returns the URL to redirect to after a successful delete.
        """
        return reverse(
            'institution_admin:student_group_list',
            kwargs={'registration_slug': self.kwargs['registration_slug']}
        )
    
    
class BusSearchFormView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, FormView):
    """
    View to handle the bus search form for a registration.
    """
    template_name = 'institution_admin/bus_search_form.html'
    form_class = BusSearchForm

    def get_registration(self):
        """
        Fetch registration using the code from the URL.
        """
        registration_code = self.kwargs.get('registration_code')
        return get_object_or_404(Registration, code=registration_code)
    
    def get_form(self, form_class=None):
        """
        Customizes the form's queryset for pickup and drop points based on registration.
        """
        form = super().get_form(form_class)
        registration = self.get_registration()
        form.fields['pickup_point'].queryset = registration.stops.all()
        form.fields['drop_point'].queryset = registration.stops.all()
        return form

    def form_valid(self, form):
        """
        Stores search criteria in the session and proceeds to the results view.
        """
        pickup_point = form.cleaned_data['pickup_point']
        drop_point = form.cleaned_data['drop_point']
        schedule = form.cleaned_data['schedule']

        # Store search criteria in the session for passing to results view
        self.request.session['pickup_point'] = pickup_point.id
        self.request.session['drop_point'] = drop_point.id
        self.request.session['schedule'] = schedule.id

        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        """
        Adds registration to the context for the template.
        """
        context = super().get_context_data(**kwargs)
        context['registration'] = get_object_or_404(Registration, code=self.kwargs.get('registration_code'))
        return context

    def get_success_url(self):
        """
        Returns the URL to redirect to after a successful form submission.
        """
        registration_code = self.get_registration().code
        ticket_id = self.kwargs.get('ticket_id')
        return reverse('institution_admin:bus_search_results', kwargs={'ticket_id': ticket_id, 'registration_code': registration_code})
    

class BusSearchResultsView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, ListView):
    """
    View to display bus search results based on session criteria.
    """
    template_name = 'institution_admin/bus_search_results.html'
    context_object_name = 'buses'

    def get_queryset(self):
        """
        Returns queryset of buses matching the search criteria from the session.
        """
        registration_code = self.kwargs.get('registration_code')
        registration = get_object_or_404(Registration, code=registration_code)

        self.pickup_point_id = self.request.session.get('pickup_point')
        self.drop_point_id = self.request.session.get('drop_point')
        self.schedule_id = self.request.session.get('schedule')

        if not (self.pickup_point_id and self.drop_point_id and self.schedule_id):
            return Bus.objects.none()
        
        buses = Bus.objects.filter(
            org=registration.org,
            schedule_id=self.schedule_id,
        ).filter(
            Q(route__stops__id=self.pickup_point_id) if self.pickup_point_id == self.drop_point_id else Q(route__stops__id__in=[self.pickup_point_id, self.drop_point_id])
        ).annotate(
            matching_stops=Count('route__stops', filter=Q(route__stops__id__in=[self.pickup_point_id, self.drop_point_id]))
        ).filter(
            matching_stops=1 if self.pickup_point_id == self.drop_point_id else 2
        ).distinct()

        # Subquery to fetch available seats from BusCapacity
        # bus_capacity_subquery = BusCapacity.objects.filter(
        #     bus=OuterRef('pk'),
        #     registration=registration
        # ).values('available_seats')

        # Annotate buses with available seats or fallback to total capacity
        # buses = buses.annotate(
        #     available_seats=Coalesce(Subquery(bus_capacity_subquery), F('capacity'))
        # )

        return buses

    def get_context_data(self, **kwargs):
        """
        Adds registration, ticket, pickup/drop points, and schedule to the context.
        """
        context = super().get_context_data(**kwargs)
        context['registration'] = get_object_or_404(Registration, code=self.kwargs.get('registration_code'))
        context['ticket'] = get_object_or_404(Ticket, ticket_id=self.kwargs.get('ticket_id'))
        context['pickup_point'] = get_object_or_404(Stop, id=self.pickup_point_id)
        context['drop_point'] = get_object_or_404(Stop, id=self.drop_point_id)
        context['schedule'] = get_object_or_404(Schedule, id=self.schedule_id)
        return context
    

class TicketExportView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, View):
    """
    View to trigger export of tickets to Excel via Celery.
    """
    def post(self, request, *args, **kwargs):
        """
        Triggers the Celery task to export tickets and returns a JSON response.
        """
        registration_slug = self.kwargs.get('registration_slug')
        search_term = request.GET.get('search', '')
        # Always filter by the current user's institution for institution admin
        filters = {
            'institution': request.user.profile.institution.slug,
            'pickup_points': request.GET.getlist('pickup_point'),
            'drop_points': request.GET.getlist('drop_point'),
            'schedule': request.GET.get('schedule'),
            'pickup_buses': request.GET.getlist('pickup_bus'),
            'drop_buses': request.GET.getlist('drop_bus'),
            'student_group': request.GET.get('student_group'),
        }

        # Trigger the Celery task
        export_tickets_to_excel.apply_async(
            args=[request.user.id, registration_slug, search_term, filters]
        )

        return JsonResponse({"message": "Export request received. You will be notified once the export is ready."})
    
    
class StopSelectFormView(FormView):
    """
    View to handle stop selection form for students.
    """
    template_name = 'students/search_form.html'
    form_class = StopSelectForm

    def get_registration(self):
        """
        Fetch registration using the code from the URL.
        """
        registration_code = self.kwargs.get('registration_code')
        return get_object_or_404(Registration, code=registration_code)
    
    def get_form(self, form_class=None):
        """
        Customizes the form's queryset for stops based on registration.
        """
        form = super().get_form(form_class)
        registration = self.get_registration()
        form.fields['stop'].queryset = registration.stops.all().order_by('name')
        return form
    
    def get_context_data(self, **kwargs):
        """
        Adds registration to the context for the template.
        """
        context = super().get_context_data(**kwargs)
        context['registration'] = get_object_or_404(Registration, code=self.kwargs.get('registration_code'))
        return context

    def form_valid(self, form):
        """
        Stores the selected stop in the session and proceeds to the next step.
        """
        stop = form.cleaned_data['stop']
        self.request.session['stop_id'] = stop.id
        return super().form_valid(form)

    def get_success_url(self):
        """
        Returns the URL to redirect to after a successful form submission.
        """
        registration_code = self.get_registration().code
        query_string = self.request.GET.get('type', '')
        return reverse('institution_admin:schedule_group_select', kwargs={'registration_code': registration_code, 'ticket_id': self.kwargs.get('ticket_id')}) + f"?type={query_string}"
    

class SelectScheduleGroupView(View):
    """
    View to handle schedule group selection for a ticket.
    """
    template_name = 'institution_admin/select_schedule_group.html'

    def get(self, request, registration_code, ticket_id):
        """
        Renders the schedule group selection page.
        """
        registration = get_object_or_404(Registration, code=registration_code)
        schedules = Schedule.objects.filter(org=registration.org, registration=registration)
        query_string = self.request.GET.get('type', '')
        type = query_string if query_string else 'pickup and drop'
        return render(request, self.template_name, {'schedules': schedules, 'type': type})

    def post(self, request, registration_code, ticket_id):
        """
        Handles the schedule group selection and stores it in the session.
        """
        selected_id = request.POST.get("schedule_group")

        if not selected_id:
            registration = get_object_or_404(Registration, code=registration_code)
            schedules = Schedule.objects.filter(org=registration.org, registration=registration)
            query_string = self.request.GET.get('type', '')
            type = query_string if query_string else 'pickup and drop'
            return render(
                request,
                self.template_name,
                {
                    'schedules': schedules,
                    'type': type,
                    'error_message': "Please select a schedule.",
                }
            )

        selected_schedule = Schedule.objects.get(id=selected_id)
        
        # Process the selection
        selection_details = {
            "selected_schedule": selected_schedule,
        }
        
        self.request.session['schedule_id'] = selection_details['selected_schedule'].id
        query_string = self.request.GET.get('type', '')
        return HttpResponseRedirect(reverse('institution_admin:bus_search_results', kwargs={'registration_code': registration_code, 'ticket_id': self.kwargs.get('ticket_id')})+ f"?type={query_string}")
    
    
class BusSearchResultsView(ListView):
    """
    View (for students) to display filtered bus records for a stop and schedule.
    """
    template_name = 'institution_admin/search_results.html'
    context_object_name = 'buses'

    def get_queryset(self):
        """
        Returns queryset of buses filtered by stop and schedule from the session.
        """
        # Get pickup point, drop point, and schedule from session
        stop_id = self.request.session.get('stop_id')
        schedule_id = self.request.session.get('schedule_id')
        query_string = self.request.GET.get('type', '')

        schedule = Schedule.objects.get(id=int(schedule_id))
        
        if query_string != '':
            schedule_ids = [schedule.id]
        else:
            raise Http404("Invalid query string")
        
        buses = get_filtered_bus_records(schedule_ids, int(stop_id))
        return buses
    
    def get(self, request, *args, **kwargs):
        """
        Redirects to 'bus_not_found' if no buses are found, otherwise renders results.
        """
        self.object_list = self.get_queryset()
        if not self.object_list:
            registration_code = self.kwargs.get('registration_code')
            return HttpResponseRedirect(reverse('students:bus_not_found', kwargs={'registration_code': registration_code}))
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        """
        Adds registration, ticket, change type, stop, and schedule to the context.
        """
        context = super().get_context_data(**kwargs)
        context['registration'] = get_object_or_404(Registration, code=self.kwargs.get('registration_code'))
        context['ticket'] = get_object_or_404(Ticket, ticket_id=self.kwargs.get('ticket_id'))
        context['change_type'] = self.request.GET.get('type', '')
        context['stop'] = Stop.objects.get(id=self.request.session.get('stop_id'))
        context['schedule'] = Schedule.objects.get(id=self.request.session.get('schedule_id'))
        return context


class UpdateBusInfoView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, View):
    """
    View to update a ticket's bus info and trip booking counts.
    """
    @transaction.atomic
    def get(self, request, registration_code, ticket_id, bus_slug):
        """
        Updates the ticket's pickup or drop bus record and adjusts trip booking counts.
        """
        registration = get_object_or_404(Registration, code=registration_code)
        ticket = get_object_or_404(Ticket, ticket_id=ticket_id)
        bus_record = get_object_or_404(BusRecord, slug=bus_slug)
        
        stop_id = self.request.session.get('stop_id')
        schedule_id = self.request.session.get('schedule_id')
        
        pickup_point = get_object_or_404(Stop, id=stop_id)
        drop_point = get_object_or_404(Stop, id=stop_id)
        schedule = get_object_or_404(Schedule, id=schedule_id)
        
        change_type = self.request.GET.get('type')
        
        if change_type == 'pickup':
            new_trip = Trip.objects.get(registration=registration, record=bus_record, schedule=schedule)
            if ticket.pickup_bus_record and ticket.pickup_schedule:
                try:
                    existing_trip = Trip.objects.get(registration=registration, record=ticket.pickup_bus_record, schedule=ticket.pickup_schedule)
                    existing_trip.booking_count -= 1
                    existing_trip.save()
                except Trip.DoesNotExist:
                    pass
            ticket.pickup_bus_record = bus_record
            ticket.pickup_point = pickup_point
            ticket.pickup_schedule = schedule
            new_trip.booking_count += 1
            new_trip.save()
        
        elif change_type == 'drop':
            new_trip = Trip.objects.get(registration=registration, record=bus_record, schedule=schedule)
            if ticket.drop_bus_record and ticket.drop_schedule:
                try:
                    existing_trip = Trip.objects.get(registration=registration, record=ticket.drop_bus_record, schedule=ticket.drop_schedule)
                    existing_trip.booking_count -= 1
                    existing_trip.save()
                except Trip.DoesNotExist:
                    pass
            ticket.drop_bus_record = bus_record
            ticket.drop_point = drop_point
            ticket.drop_schedule = schedule
            new_trip.booking_count += 1
            new_trip.save()
        
        # Update the ticket type to 'two way' if both pickup and drop buses exist
        if ticket.pickup_bus_record and ticket.drop_bus_record:
            ticket.ticket_type = 'two_way'
        
        ticket.save()
        
        return redirect(
            reverse('institution_admin:ticket_list', 
                    kwargs={'registration_slug': registration.slug}
                )
            )


class BusRequestListView(ListView):
    """
    View to list all bus requests for a registration.
    """
    model = BusRequest
    template_name = 'institution_admin/bus_request_list.html'
    context_object_name = 'bus_requests'
    paginate_by = 20
    
    def get_queryset(self):
        """
        Returns queryset of bus requests filtered by registration and institution, with optional search.
        """
        registration = get_object_or_404(Registration, slug=self.kwargs["registration_slug"])
        institution = self.request.user.profile.institution
        queryset = BusRequest.objects.filter(org=self.request.user.profile.org, institution=institution, registration=registration).order_by('-created_at')
        search_query = self.request.GET.get('search', '').strip()
        if search_query:
            queryset = queryset.filter(
                Q(student_name__icontains=search_query) |
                Q(contact_no__icontains=search_query) |
                Q(contact_email__icontains=search_query) |
                Q(receipt__receipt_id__icontains=search_query)
            )
        return queryset
    
    def get_context_data(self, **kwargs):
        """
        Adds registration and request counts to the context, and checks if each request has a ticket.
        """
        context = super().get_context_data(**kwargs)
        registration = Registration.objects.get(slug=self.kwargs["registration_slug"])
        context["registration"] = registration
        context["total_requests"] = BusRequest.objects.filter(
            org=self.request.user.profile.org, 
            institution=self.request.user.profile.institution, 
            registration=registration
        ).count()
        context["open_requests"] = BusRequest.objects.filter(
            org=self.request.user.profile.org, 
            institution=self.request.user.profile.institution, 
            registration=registration, 
            status='open'
        ).count()
        context["closed_requests"] = BusRequest.objects.filter(
            org=self.request.user.profile.org, 
            institution=self.request.user.profile.institution, 
            registration=registration, 
            status='closed'
        ).count()
        for request in context["bus_requests"]:
            request.has_ticket = Ticket.objects.filter(
                registration=registration, 
                recipt=request.receipt,
                is_terminated=False
            ).exists()
        return context

class BusRequestOpenListView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, ListView):
    """
    View to list all open bus requests for a registration.
    """
    model = BusRequest
    template_name = 'institution_admin/bus_request_list.html'
    context_object_name = 'bus_requests'
    paginate_by = 20
    
    def get_queryset(self):
        """
        Returns queryset of open bus requests filtered by registration and institution.
        """
        registration = get_object_or_404(Registration, slug=self.kwargs["registration_slug"])
        institution = self.request.user.profile.institution
        queryset = BusRequest.objects.filter(org=self.request.user.profile.org, institution=institution, registration=registration, status='open').order_by('-created_at')
        return queryset
    
    def get_context_data(self, **kwargs):
        """
        Adds registration and request counts to the context, and checks if each request has a ticket.
        """
        context = super().get_context_data(**kwargs)
        registration = Registration.objects.get(slug=self.kwargs["registration_slug"])
        context["registration"] = registration
        context["total_requests"] = BusRequest.objects.filter(
            org=self.request.user.profile.org, 
            institution=self.request.user.profile.institution, 
            registration=registration
        ).count()
        context["open_requests"] = BusRequest.objects.filter(
            org=self.request.user.profile.org, 
            institution=self.request.user.profile.institution, 
            registration=registration, 
            status='open'
        ).count()
        context["closed_requests"] = BusRequest.objects.filter(
            org=self.request.user.profile.org, 
            institution=self.request.user.profile.institution, 
            registration=registration, 
            status='closed'
        ).count()
        for request in context["bus_requests"]:
            request.has_ticket = Ticket.objects.filter(
                registration=registration, 
                recipt=request.receipt,
                is_terminated=False
            ).exists()
        return context

class BusRequestClosedListView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, ListView):
    """
    View to list all closed bus requests for a registration.
    """
    model = BusRequest
    template_name = 'institution_admin/bus_request_list.html'
    context_object_name = 'bus_requests'
    paginate_by = 20
    
    def get_queryset(self):
        """
        Returns queryset of closed bus requests filtered by registration and institution.
        """
        registration = get_object_or_404(Registration, slug=self.kwargs["registration_slug"])
        institution = self.request.user.profile.institution
        queryset = BusRequest.objects.filter(org=self.request.user.profile.org, institution=institution, registration=registration, status='closed').order_by('-created_at')
        return queryset
    
    def get_context_data(self, **kwargs):
        """
        Adds registration and request counts to the context, and checks if each request has a ticket.
        """
        context = super().get_context_data(**kwargs)
        registration = Registration.objects.get(slug=self.kwargs["registration_slug"])
        context["registration"] = registration
        context["total_requests"] = BusRequest.objects.filter(
            org=self.request.user.profile.org, 
            institution=self.request.user.profile.institution, 
            registration=registration
        ).count()
        context["open_requests"] = BusRequest.objects.filter(
            org=self.request.user.profile.org, 
            institution=self.request.user.profile.institution, 
            registration=registration, 
            status='open'
        ).count()
        context["closed_requests"] = BusRequest.objects.filter(
            org=self.request.user.profile.org, 
            institution=self.request.user.profile.institution, 
            registration=registration, 
            status='closed'
        ).count()
        for request in context["bus_requests"]:
            request.has_ticket = Ticket.objects.filter(
                registration=registration, 
                recipt=request.receipt,
                is_terminated=False
            ).exists()
        return context

class BusRequestDeleteView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, DeleteView):
    """
    View to delete a bus request.
    """
    model = BusRequest
    template_name = 'institution_admin/bus_request_confirm_delete.html'
    slug_field = 'slug'
    slug_url_kwarg = 'bus_request_slug'
    
    def get_context_data(self, **kwargs):
        """
        Adds registration to the context for the template.
        """
        context = super().get_context_data(**kwargs)
        context['registration'] = self.object.registration
        return context
    
    def get_success_url(self):
        """
        Returns the URL to redirect to after a successful delete.
        """
        return reverse('central_admin:bus_request_list', kwargs={'registration_slug': self.kwargs['registration_slug']})


class BusRequestStatusUpdateView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, View):
    """
    View to update the status of a bus request and add comments.
    """
    def post(self, request, *args, **kwargs):
        """
        Toggles the status of a bus request and adds a comment if provided.
        """
        bus_request = get_object_or_404(BusRequest, slug=self.kwargs['bus_request_slug'])
        new_status = 'open' if bus_request.status == 'closed' else 'closed'
        comment_text = request.POST.get('comment')
        bus_request.status = new_status
        bus_request.save()
        if comment_text:
            BusRequestComment.objects.create(
                bus_request=bus_request,
                comment=comment_text,
                created_by=request.user
            )
        modal_body_html = render_to_string('central_admin/bus_request_modal_body.html', {'bus_request': bus_request})
        response = HttpResponse(modal_body_html)
        response['HX-Trigger'] = 'reloadPage'
        return response

class BusRequestCommentView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, View):
    """
    View to add a comment to a bus request.
    """
    def post(self, request, *args, **kwargs):
        """
        Handles the submission of a comment form for a bus request.
        """
        bus_request = get_object_or_404(BusRequest, slug=self.kwargs['bus_request_slug'])
        comment_form = BusRequestCommentForm(request.POST)
        if comment_form.is_valid():
            comment = BusRequestComment.objects.create(
                bus_request=bus_request,
                comment=comment_form.cleaned_data['comment'],
                created_by=request.user
            )
            comment_html = render_to_string('institution_admin/comment.html', {'comment': comment}).strip()
            return HttpResponse(comment_html)
        return HttpResponse('Invalid form submission', status=400)


class BulkStudentGroupUpdateView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, FormView):
    """
    View to handle bulk update of student groups via Excel upload.
    """
    template_name = 'institution_admin/bulk_student_group_update_upload.html'
    form_class = BulkStudentGroupUpdateForm

    def get(self, request, *args, **kwargs):
        """
        Renders the upload form for bulk student group update.
        """
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def form_valid(self, form):
        """
        Validates and previews the bulk update data from the uploaded Excel file.
        """
        file = form.cleaned_data['file']
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        preview_data = []
        errors = []
        institution = self.request.user.profile.institution

        # Validate header row
        header = [str(cell).strip().upper() if cell else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        expected_header = ["STUDENT ID", "CLASS", "SECTION"]
        if header != expected_header:
            messages.error(self.request, "Excel file must have only these columns in order: STUDENT ID, CLASS, SECTION (no extra columns or data).")
            return render(self.request, self.template_name, {
                "form": form,
                "errors": errors,
            })

        # Collect all student_ids from the file for bulk query
        student_id_rows = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or len(row) != 3:
                errors.append(f"Row {idx}: Must have exactly 3 columns (STUDENT ID, CLASS, SECTION).")
                continue
            student_id, class_name, section = row
            if not student_id or not class_name or not section:
                errors.append(f"Row {idx}: Missing data")
                continue
            student_id_rows.append((idx, student_id, class_name, section))

        # Bulk fetch all tickets for student_ids
        student_ids = [str(student_id).strip() for _, student_id, _, _ in student_id_rows]
        tickets = Ticket.objects.filter(student_id__in=student_ids, institution=institution).select_related('student_group')
        ticket_map = {t.student_id: t for t in tickets}

        for idx, student_id, class_name, section in student_id_rows:
            group_name = f"{str(class_name).strip().upper()} - {str(section).strip().upper()}"
            ticket = ticket_map.get(str(student_id).strip())
            if ticket:
                current_group = ticket.student_group.name if ticket.student_group else ""
                preview_data.append({
                    "student_id": student_id,
                    "student_name": ticket.student_name,
                    "current_group": current_group,
                    "new_group": group_name,
                })
            else:
                errors.append(f"Row {idx}: Ticket not found for student_id {student_id}")

        self.request.session['bulk_update_preview'] = preview_data
        self.request.session['bulk_update_errors'] = errors
        return render(self.request, 'institution_admin/bulk_student_group_update_confirm.html', {
            "preview_data": preview_data,
            "errors": errors,
        })

    def dispatch(self, request, *args, **kwargs):
        """
        Ensures bulk update is only allowed when registration is closed.
        """
        registration_slug = self.kwargs.get('registration_slug')
        registration = get_object_or_404(Registration, slug=registration_slug)
        if registration.status:
            raise Http404("Bulk update is only allowed when registration is closed.")
        return super().dispatch(request, *args, **kwargs)

class BulkStudentGroupUpdateConfirmView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, View):
    """
    View to confirm and process bulk student group updates.
    """
    def post(self, request, *args, **kwargs):
        """
        Triggers the Celery task to process the bulk update and clears the preview session.
        """
        preview_data = request.session.get('bulk_update_preview', [])
        institution = request.user.profile.institution
        user_id = request.user.id
        # Use Celery for background processing to avoid blocking the request
        bulk_update_student_groups_task.delay(user_id, institution.id, preview_data)
        request.session.pop('bulk_update_preview', None)
        return HttpResponseRedirect(
            reverse('institution_admin:ticket_list', kwargs={'registration_slug': self.kwargs['registration_slug']})
        )
    
    def dispatch(self, request, *args, **kwargs):
        """
        Ensures bulk update is only allowed when registration is closed.
        """
        registration_slug = self.kwargs.get('registration_slug')
        registration = get_object_or_404(Registration, slug=registration_slug)
        if registration.status:
            raise Http404("Bulk update is only allowed when registration is closed.")
        return super().dispatch(request, *args, **kwargs)


class ReservationListView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, ListView):
    """
    ReservationListView displays the list of reservations for institution admin users.
    Only shows reservations created by their institution.
    
    This view inherits from:
        - LoginRequiredMixin: Ensures that the user is authenticated.
        - InsitutionAdminOnlyAccessMixin: Ensures that the user has institution admin access.
        - ListView: Provides list display functionality.
    
    Attributes:
        model (BusReservationRequest): The model to list.
        template_name (str): The template to render for this view.
        context_object_name (str): The context variable name for the list.
    """
    model = BusReservationRequest
    template_name = "institution_admin/reservation_list.html"
    context_object_name = 'reservations'
    
    def get_queryset(self):
        """
        Returns queryset of reservations filtered by the institution.
        """
        return BusReservationRequest.objects.filter(
            institution=self.request.user.profile.institution
        ).select_related('institution', 'created_by').order_by('-created_at')


class ReservationDetailView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, DetailView):
    """
    ReservationDetailView displays the details of a specific reservation for institution admin users.
    """
    model = BusReservationRequest
    template_name = "institution_admin/reservation_detail.html"
    context_object_name = 'reservation'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    
    def get_queryset(self):
        """
        Ensures institution admins can only view their own institution's reservations.
        """
        return BusReservationRequest.objects.filter(
            institution=self.request.user.profile.institution
        )


class ReservationCreateView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, CreateView):
    """
    ReservationCreateView allows institution admins to create new bus reservation requests.
    
    This view inherits from:
        - LoginRequiredMixin: Ensures that the user is authenticated.
        - InsitutionAdminOnlyAccessMixin: Ensures that the user has institution admin access.
        - CreateView: Provides create functionality for the model.
    
    Attributes:
        model (BusReservationRequest): The model to create.
        template_name (str): The template to render for this view.
        form_class (BusReservationRequestForm): The form class for the reservation.
    """
    model = BusReservationRequest
    template_name = "institution_admin/reservation_create.html"
    form_class = BusReservationRequestForm
    
    def form_valid(self, form):
        """
        Saves the new reservation request with org, institution, and created_by set automatically.
        Extracts date and time from datetime fields and calculates duration.
        """
        reservation = form.save(commit=False)
        reservation.org = self.request.user.profile.org
        reservation.institution = self.request.user.profile.institution
        reservation.created_by = self.request.user
        reservation.status = 'pending'
        
        # Extract date and time from datetime fields
        departure_datetime = form.cleaned_data['departure_datetime']
        arrival_datetime = form.cleaned_data['arrival_datetime']
        
        reservation.date = departure_datetime.date()
        reservation.departure_time = departure_datetime.time()
        reservation.arrival_time = arrival_datetime.time()
        
        # Calculate total duration
        reservation.total_duration = arrival_datetime - departure_datetime
        
        reservation.save()
        messages.success(self.request, "Bus reservation request created successfully!")
        return redirect('institution_admin:reservation_list')
    
    def get_context_data(self, **kwargs):
        """
        Adds additional context to the template.
        """
        context = super().get_context_data(**kwargs)
        return context


class ReservationDeleteView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, DeleteView):
    """
    ReservationDeleteView allows institution admins to delete their pending reservation requests.
    
    This view inherits from:
        - LoginRequiredMixin: Ensures that the user is authenticated.
        - InsitutionAdminOnlyAccessMixin: Ensures that the user has institution admin access.
        - DeleteView: Provides delete functionality for the model.
    
    Attributes:
        model (BusReservationRequest): The model to delete.
        slug_field (str): The field used for slug lookup.
        slug_url_kwarg (str): The URL keyword argument for the slug.
        success_url (str): URL to redirect to after successful deletion.
    """
    model = BusReservationRequest
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('institution_admin:reservation_list')
    
    def get_queryset(self):
        """
        Ensures institution admins can only delete their own institution's pending reservations.
        """
        return BusReservationRequest.objects.filter(
            institution=self.request.user.profile.institution,
            status='pending'  # Only allow deletion of pending reservations
        )
    
    def post(self, request, *args, **kwargs):
        """
        Handles the deletion and shows success message.
        """
        messages.success(request, "Reservation deleted successfully!")
        return super().delete(request, *args, **kwargs)


class PaymentListView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, ListView):
    """
    View to list all payments for the current user's institution.
    
    Attributes:
        model (Payment): The Payment model to query.
        template_name (str): Template for displaying payments list.
        context_object_name (str): Context variable name for payments.
        paginate_by (int): Number of payments per page.
    """
    model = None  # Will be imported dynamically
    template_name = 'institution_admin/payment_list.html'
    context_object_name = 'payments'
    paginate_by = 50
    
    def dispatch(self, request, *args, **kwargs):
        # Import Payment model here to avoid circular import
        from services.models import Payment
        self.model = Payment
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        """
        Returns payments filtered by institution and registration.
        """
        queryset = self.model.objects.filter(
            institution=self.request.user.profile.institution
        ).select_related('ticket', 'registration', 'installment_date', 'recorded_by')
        
        # Filter by registration if provided
        registration_slug = self.kwargs.get('registration_slug')
        if registration_slug:
            queryset = queryset.filter(registration__slug=registration_slug)
        
        # Filter by ticket if provided in query params
        ticket_slug = self.request.GET.get('ticket')
        if ticket_slug:
            queryset = queryset.filter(ticket__slug=ticket_slug)
        
        return queryset.order_by('-payment_date', '-created_at')
    
    def get_context_data(self, **kwargs):
        """
        Adds registration info to context.
        """
        context = super().get_context_data(**kwargs)
        registration_slug = self.kwargs.get('registration_slug')
        if registration_slug:
            context['registration'] = get_object_or_404(
                Registration,
                slug=registration_slug,
                org=self.request.user.profile.org
            )
        return context


class PaymentCreateView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, CreateView):
    """
    View to record a new payment for a ticket.
    
    Attributes:
        model (Payment): The Payment model.
        template_name (str): Template for payment creation form.
        form_class: PaymentForm from institution_admin forms.
    """
    model = None  # Will be imported dynamically
    template_name = 'institution_admin/payment_form.html'
    form_class = None  # Will be set in dispatch
    
    def dispatch(self, request, *args, **kwargs):
        # Import Payment model and form here to avoid circular import
        from services.models import Payment
        from services.forms.institution_admin import PaymentForm
        self.model = Payment
        self.form_class = PaymentForm
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        """
        Passes institution, registration, and ticket to the form.
        """
        kwargs = super().get_form_kwargs()
        kwargs['institution'] = self.request.user.profile.institution
        
        registration_slug = self.kwargs.get('registration_slug')
        if registration_slug:
            registration = get_object_or_404(
                Registration,
                slug=registration_slug,
                org=self.request.user.profile.org
            )
            kwargs['registration'] = registration
        
        # Get ticket from URL and pass to form for validation
        ticket_slug = self.request.GET.get('ticket') or self.request.POST.get('ticket_slug')
        if ticket_slug:
            try:
                from services.models import Ticket
                ticket = Ticket.objects.get(
                    slug=ticket_slug,
                    institution=self.request.user.profile.institution
                )
                kwargs['ticket'] = ticket
            except Ticket.DoesNotExist:
                pass

        
        return kwargs
    
    def get_initial(self):
        """
        Ticket is no longer in the form - it will be set from URL in form_valid.
        """
        initial = super().get_initial()
        return initial
    
    def form_valid(self, form):
        """
        Sets the ticket, org, registration, institution, and recorded_by before saving.
        Ticket is obtained from URL parameter.
        """
        payment = form.save(commit=False)
        
        # Get ticket from URL parameter
        ticket_slug = self.request.GET.get('ticket') or self.request.POST.get('ticket_slug')
        if not ticket_slug:
            messages.error(self.request, "Ticket parameter is required.")
            return self.form_invalid(form)
            
        try:
            payment.ticket = Ticket.objects.get(
                slug=ticket_slug,
                institution=self.request.user.profile.institution
            )
        except Ticket.DoesNotExist:
            messages.error(self.request, "Invalid ticket specified.")
            return self.form_invalid(form)
        
        payment.org = self.request.user.profile.org
        payment.institution = self.request.user.profile.institution
        payment.registration = payment.ticket.registration
        payment.recorded_by = self.request.user
        payment.save()
        
        messages.success(self.request, f"Payment {payment.payment_id} recorded successfully!")
        return redirect('institution_admin:payment_list', registration_slug=payment.registration.slug)
    
    def get_context_data(self, **kwargs):
        """
        Adds registration info to context.
        """
        context = super().get_context_data(**kwargs)
        registration_slug = self.kwargs.get('registration_slug')
        if registration_slug:
            context['registration'] = get_object_or_404(
                Registration,
                slug=registration_slug,
                org=self.request.user.profile.org
            )
        
        # Pass ticket to context if provided in URL
        ticket_slug = self.request.GET.get('ticket')
        if ticket_slug:
            try:
                ticket = Ticket.objects.get(
                    slug=ticket_slug,
                    institution=self.request.user.profile.institution
                )
                context['ticket'] = ticket
            except Ticket.DoesNotExist:
                pass
        
        return context


class PaymentDetailView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, DetailView):
    """
    View to display details of a single payment.
    
    Attributes:
        model (Payment): The Payment model.
        template_name (str): Template for payment details.
        context_object_name (str): Context variable name for payment.
        slug_field (str): The field used for slug lookup.
        slug_url_kwarg (str): The URL keyword argument for the slug.
    """
    model = None  # Will be imported dynamically
    template_name = 'institution_admin/payment_detail.html'
    context_object_name = 'payment'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    
    def dispatch(self, request, *args, **kwargs):
        # Import Payment model here to avoid circular import
        from services.models import Payment
        self.model = Payment
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        """
        Ensures institution admins can only view their institution's payments.
        """
        return self.model.objects.filter(
            institution=self.request.user.profile.institution
        ).select_related('ticket', 'registration', 'installment_date', 'recorded_by')


class PaymentDeleteView(LoginRequiredMixin, InsitutionAdminOnlyAccessMixin, DeleteView):
    """
    View to delete a payment record.
    
    Attributes:
        model (Payment): The Payment model.
        slug_field (str): The field used for slug lookup.
        slug_url_kwarg (str): The URL keyword argument for the slug.
    """
    model = None  # Will be imported dynamically
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    
    def dispatch(self, request, *args, **kwargs):
        # Import Payment model here to avoid circular import
        from services.models import Payment
        self.model = Payment
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        """
        Ensures institution admins can only delete their institution's payments.
        """
        return self.model.objects.filter(
            institution=self.request.user.profile.institution
        )
    
    def get_success_url(self):
        """
        Redirects to payment list after successful deletion.
        """
        registration_slug = self.object.registration.slug
        messages.success(self.request, f"Payment {self.object.payment_id} deleted successfully!")
        return reverse('institution_admin:payment_list', kwargs={'registration_slug': registration_slug})

